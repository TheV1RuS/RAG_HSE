from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.config import settings
from src.debug_dump import append_debug_record
from src.rag_answer import answer_question
from src.utils import ensure_parent


def _find_column(headers: dict[str, int], preferred: str) -> int | None:
    low = preferred.lower().strip()
    for name, idx in headers.items():
        if name.lower().strip() == low:
            return idx
    return None


def fill_testset(input_path: str | Path, output_path: str | Path) -> Path:
    wb = load_workbook(filename=str(input_path))
    ws = wb.active

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            headers[str(value)] = col

    q_col = _find_column(headers, "question")
    if q_col is None:
        q_col = 1
        print("[WARN] Column 'question' not found, fallback to first column.")

    a_col = _find_column(headers, "answer")
    if a_col is None:
        a_col = ws.max_column + 1
        ws.cell(row=1, column=a_col, value="answer")

    for row in range(2, ws.max_row + 1):
        question = ws.cell(row=row, column=q_col).value
        if question is None or not str(question).strip():
            continue
        answer, debug = answer_question(str(question).strip())
        ws.cell(row=row, column=a_col, value=answer)
        append_debug_record(settings.debug_dump_path, debug)

    out = ensure_parent(output_path)
    wb.save(str(out))
    return out
